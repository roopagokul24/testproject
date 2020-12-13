import datetime
import openpyxl

from django.db.models import Count,Func, F, Value
from django.db.models.functions import Sqrt
from django.utils import timezone
from django.conf import settings
# from dateutil import parser


from django.db.models import Q
from django.shortcuts import render
from django.template.loader import render_to_string
from django.views.generic import TemplateView,View, CreateView
from django.views.generic.edit import FormView
from django.urls import reverse, reverse_lazy
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.http import HttpResponseRedirect
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from apps.dashboard.forms import UploadDataForm

from apps.matches.models import Team, Match, Player, Umpire, Venue
from django.contrib.sessions.models import Session


class DashboardHomeView(TemplateView):
    template_name = 'dashboard/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Session.objects.all().delete()
        return context


class TeamView(TemplateView):
    template_name = 'dashboard/team.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        list = Team.objects.all()
        context['list'] = list
        return context

class PlayerView(TemplateView):
    template_name = 'dashboard/player.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        list = Player.objects.all()
        context['list'] = list
        return context

class VenueView(TemplateView):
    template_name = 'dashboard/venue.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        list = Venue.objects.all()
        context['list'] = list
        return context

class UmpireView(TemplateView):
    template_name = 'dashboard/umpire.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        list = Umpire.objects.all()
        context['list'] = list
        return context

class MatchView(TemplateView):
    template_name = 'dashboard/match.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        list = Match.objects.all()
        context['list'] = list
        return context


class UploadTeamView(FormView):
    template_name = 'dashboard/upload-data.html'
    form_class = UploadDataForm
    success_url = reverse_lazy('dashboard:teams')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('file')
        # print("file:",excel_file)

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        num_of_rows = len(excel_data)
        for i in range(1,num_of_rows+1):
            index = "A"+str(i)
            name = worksheet[index].value
            obj = Team()
            obj.name = name
            obj.save()
        messages.success(self.request,'Teams added successfully.')
        return super().form_valid(form)

    def form_invalid(self, form):
        print(form.errors)
        messages.warning(self.request, 'Form Submission Failed !!')
        return super().form_invalid(form)

class UploadPlayerView(FormView):
    template_name = 'dashboard/upload-data.html'
    form_class = UploadDataForm
    success_url = reverse_lazy('dashboard:players')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('file')
        # print("file:",excel_file)

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        num_of_rows = len(excel_data)
        for i in range(1,num_of_rows+1):
            index = "A"+str(i)
            name = worksheet[index].value
            obj = Player()
            obj.name = name
            obj.save()
        messages.success(self.request,'Players added successfully.')
        return super().form_valid(form)

    def form_invalid(self, form):
        print(form.errors)
        messages.warning(self.request, 'Form Submission Failed !!')
        return super().form_invalid(form)


class UploadVenueView(FormView):
    template_name = 'dashboard/upload-data.html'
    form_class = UploadDataForm
    success_url = reverse_lazy('dashboard:venues')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('file')
        # print("file:",excel_file)

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        num_of_rows = len(excel_data)
        for i in range(1,num_of_rows+1):
            index = "A"+str(i)
            name = worksheet[index].value
            obj = Venue()
            obj.name = name
            obj.save()
        messages.success(self.request,'Venues added successfully.')
        return super().form_valid(form)

    def form_invalid(self, form):
        print(form.errors)
        messages.warning(self.request, 'Form Submission Failed !!')
        return super().form_invalid(form)


class UploadUmpireView(FormView):
    template_name = 'dashboard/upload-data.html'
    form_class = UploadDataForm
    success_url = reverse_lazy('dashboard:umpires')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('file')
        # print("file:",excel_file)

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        num_of_rows = len(excel_data)
        for i in range(1,num_of_rows+1):
            index = "A"+str(i)
            name = worksheet[index].value
            obj = Umpire()
            obj.name = name
            obj.save()
        messages.success(self.request,'Umpire added successfully.')
        return super().form_valid(form)

    def form_invalid(self, form):
        print(form.errors)
        messages.warning(self.request, 'Form Submission Failed !!')
        return super().form_invalid(form)


class UploadMatchView(FormView):
    template_name = 'dashboard/upload-data.html'
    form_class = UploadDataForm
    success_url = reverse_lazy('dashboard:matches')

    def form_valid(self, form):
        excel_file = form.cleaned_data.get('file')
        # print("file:",excel_file)

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]
        # print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        num_of_rows = len(excel_data)
        for i in range(1,num_of_rows+1):

            index = "A"+str(i)
            season = worksheet[index].value

            index = "B"+str(i)
            city = worksheet[index].value

            # index = "C"+str(i)
            # date = worksheet[index].value

            index = "D"+str(i)
            team1 = worksheet[index].value

            index = "E"+str(i)
            team2 = worksheet[index].value

            index = "F"+str(i)
            toss_winner = worksheet[index].value

            index = "G"+str(i)
            toss_decision = worksheet[index].value

            index = "H"+str(i)
            result = worksheet[index].value

            index = "I"+str(i)
            dl_applied = worksheet[index].value

            index = "J"+str(i)
            winner = worksheet[index].value

            index = "K"+str(i)
            win_by_runs = worksheet[index].value

            index = "L"+str(i)
            win_by_wickets = worksheet[index].value

            index = "M"+str(i)
            player_of_match = worksheet[index].value

            index = "N"+str(i)
            venue = worksheet[index].value

            index = "O"+str(i)
            umpire1 = worksheet[index].value

            index = "P"+str(i)
            umpire2 = worksheet[index].value

            index = "Q"+str(i)
            umpire3 = worksheet[index].value

            obj = Match()
            obj.season = season
            obj.city = city
            # obj.date = date
            obj.team1 =  Team.objects.filter(name=team1).first()
            obj.team2 =  Team.objects.filter(name=team2).first()
            obj.toss_winner =  Team.objects.filter(name=toss_winner).first()

            if toss_decision == 'field':
                obj.toss_decision =  Match.FIELD
            elif toss_decision == 'bat':
                obj.toss_decision =  Match.BAT

            if result == 'normal':
                obj.result =  Match.NORMAL
            elif result == 'tie':
                obj.result =  Match.TIE
            if result == 'no result':
                obj.result =  Match.NORESULT
            if dl_applied == None:
                obj.dl_applied =  0
            else:
                obj.dl_applied =  dl_applied

            obj.winner =  Team.objects.filter(name=winner).first()

            if win_by_runs == None:
                obj.win_by_runs =  0
            else:
                obj.win_by_runs =  win_by_runs

            obj.win_by_runs =  win_by_wickets
            obj.player_of_match =  Player.objects.filter(name=player_of_match).first()
            obj.venue =  Venue.objects.filter(name=venue).first()
            obj.umpire1 =  Venue.objects.filter(name=umpire1).first()
            obj.umpire2 =  Venue.objects.filter(name=umpire2).first()
            obj.umpire3 =  Venue.objects.filter(name=umpire3).first()

            obj.save()
        messages.success(self.request,'Match added successfully.')
        return super().form_valid(form)

    def form_invalid(self, form):
        print(form.errors)
        messages.warning(self.request, 'Form Submission Failed !!')
        return super().form_invalid(form)
